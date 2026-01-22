#!/usr/bin/env python3
"""
Screenshot Validation Module for SAP Daily Checks Audit

Extracts embedded images from Excel files and uses Azure OpenAI GPT-5.1
to validate that screenshot content matches reported check values.
"""

import base64
import json
import os
import re
from pathlib import Path
from dataclasses import dataclass
from typing import Any, List, Optional
from pydantic import BaseModel, Field

try:
    from openai import AzureOpenAI
    HAS_AZURE = True
except ImportError:
    HAS_AZURE = False

try:
    import openpyxl
    from openpyxl.drawing.image import Image as OpenpyxlImage
except ImportError:
    pass

# Load .env file if present
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    # Try manual loading if dotenv not installed
    env_path = Path(__file__).parent.parent.parent.parent.parent / '.env'
    if env_path.exists():
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    os.environ[key.strip()] = value.strip()


# --- Pydantic Models for Structured Output ---

class ExtractedScreenshotData(BaseModel):
    failed_data_backup: Optional[int] = Field(None, description="Number of failed data backups visible")
    failed_log_backup: Optional[int] = Field(None, description="Number of failed log backups visible")
    failed_jobs: Optional[int] = Field(None, description="Number of failed jobs visible")
    successful_backups: Optional[int] = Field(None, description="Number of successful backups visible")
    total_entries: Optional[int] = Field(None, description="Total number of entries/rows visible in the table")
    has_errors: bool = Field(False, description="Whether any error indicators (red icons, 'failed' status) are present")
    error_indicators: List[str] = Field(default_factory=list, description="List of specific error messages or red status text strings found")
    
    class Config:
        extra = 'forbid'

class ScreenshotAnalysisResponse(BaseModel):
    type: str = Field(..., description="Type of data shown: 'backup', 'jobs', 'logs', or 'other'")
    summary: str = Field(..., description="Brief description of what is seen in the screenshot")
    data: ExtractedScreenshotData
    
    class Config:
        extra = 'forbid'


# --- Internal Data Structures ---

@dataclass
class ScreenshotAnalysis:
    """Results from analyzing a screenshot."""
    image_name: str
    sheet_name: str
    analysis_type: str
    extracted_data: dict
    raw_response: str


@dataclass
class ValidationIssue:
    """A discrepancy between screenshot and reported values."""
    sheet: str
    image_name: str
    severity: str  # 'critical', 'warning'
    message: str
    screenshot_value: Any
    reported_value: Any


class ScreenshotValidator:
    """Validates Excel screenshots against reported check values."""
    
    VISION_PROMPT = """You are an SAP audit vision expert.
ANALYZE the image to find the 'Failed data backup' and 'Failed log backup' counts.
Also look for 'failed jobs' or 'updates'.
The numbers are usually in Green (OK) or Red (Failed) cells.
You MUST extract these integers.
If the image shows a table, count the rows or find summary numbers.
If the image is unrelated, classify as 'other'."""

    def __init__(self, workbook_path: str):
        self.workbook_path = Path(workbook_path)
        self.workbook = None
        self.extracted_images: list[tuple[str, str, bytes]] = []
        self.analyses: list[ScreenshotAnalysis] = []
        self.validation_issues: list[ValidationIssue] = []
        
        # Initialize Azure Client
        self.azure_client = None
        self.use_azure = False
        
        azure_key = os.environ.get('AZURE_OPENAI_API_KEY')
        azure_endpoint = os.environ.get('AZURE_OPENAI_ENDPOINT')
        
        if HAS_AZURE and azure_key and azure_endpoint:
            self.use_azure = True
            self.azure_client = AzureOpenAI(
                api_key=azure_key,
                api_version=os.environ.get('AZURE_OPENAI_API_VERSION', '2024-08-01-preview'),
                azure_endpoint=azure_endpoint
            )
            self.azure_deployment = os.environ.get('AZURE_OPENAI_DEPLOYMENT_NAME', 'gpt-5.1')
    
    def extract_images_from_excel(self) -> list[tuple[str, str, bytes]]:
        """Extract all embedded images from the Excel workbook."""
        self.workbook = openpyxl.load_workbook(self.workbook_path)
        images = []
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            if hasattr(sheet, '_images'):
                for idx, img in enumerate(sheet._images):
                    try:
                        image_data = img._data()
                        image_id = f"{sheet_name}_img_{idx}"
                        images.append((sheet_name, image_id, image_data))
                    except Exception as e:
                        print(f"Warning: Could not extract image {idx} from {sheet_name}: {e}")
        
        self.extracted_images = images
        return images
    
    def analyze_image_with_azure(self, image_bytes: bytes, image_name: str) -> dict | None:
        """Use Azure OpenAI with Structured Outputs via beta.parse()."""
        if not self.azure_client:
            return None

        try:
            image_b64 = base64.standard_b64encode(image_bytes).decode('utf-8')
            mime_type = "image/png"
            if image_bytes[:2] == b'\xff\xd8':
                mime_type = "image/jpeg"

            # Use the new beta parse method which handles Pydantic schemas automatically
            completion = self.azure_client.beta.chat.completions.parse(
                model=self.azure_deployment,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": self.VISION_PROMPT},
                            {
                                "type": "image_url",
                                "image_url": {"url": f"data:{mime_type};base64,{image_b64}"}
                            }
                        ]
                    }
                ],
                response_format=ScreenshotAnalysisResponse
            )
            
            message = completion.choices[0].message
            
            # If the model refused to parse, it might be a refusal
            if message.refusal:
                print(f"Refusal for {image_name}: {message.refusal}")
                return None
            
            if message.parsed:
                return message.parsed.model_dump()
                
            return None

        except Exception as e:
            print(f"Error analyzing image {image_name} with Azure: {e}")
            return None

    def analyze_all_images(self) -> list[ScreenshotAnalysis]:
        """Analyze all extracted images."""
        analyses = []
        
        for sheet_name, image_id, image_bytes in self.extracted_images:
            print(f"  Analyzing {image_id}...")
            
            result = None
            if self.use_azure:
                result = self.analyze_image_with_azure(image_bytes, image_id)
            
            if result is not None:
                analysis = ScreenshotAnalysis(
                    image_name=image_id,
                    sheet_name=sheet_name,
                    analysis_type=result.get('type', 'unknown'),
                    extracted_data=result.get('data', {}),
                    raw_response=result.get('summary', '')
                )
                analyses.append(analysis)
        
        self.analyses = analyses
        return analyses
    
    def extract_reported_values(self, sheet_name: str) -> dict:
        """Extract the reported check values from a sheet."""
        sheet = self.workbook[sheet_name]
        values = {}
        failed_jobs_total = 0
        
        for row in sheet.iter_rows(min_row=1, values_only=True):
            row_text = ' '.join(str(cell) for cell in row if cell).lower()
            
            if 'failed data backup' in row_text:
                for cell in row:
                    if isinstance(cell, (int, float)):
                        values['failed_data_backup'] = int(cell)
                        break
            
            if 'failed log backup' in row_text:
                for cell in row:
                    if isinstance(cell, (int, float)):
                        values['failed_log_backup'] = int(cell)
                        break
            
            # Handle both "Number of Failed Jobs: Today/Yesterday" and "failed jobs no." patterns
            if ('failed' in row_text and 'job' in row_text) or ('number of failed jobs' in row_text):
                for cell in row:
                    if isinstance(cell, (int, float)):
                        failed_jobs_total += int(cell)
                        break
        
        if failed_jobs_total > 0:
            values['failed_jobs'] = failed_jobs_total
        
        return values

    
    def validate_against_reports(self) -> list[ValidationIssue]:
        """Compare screenshot analyses with reported values."""
        issues = []
        
        for analysis in self.analyses:
            if analysis.analysis_type == 'unknown':
                continue
            
            reported = self.extract_reported_values(analysis.sheet_name)
            extracted = analysis.extracted_data
            
            # Check failed data backup
            if extracted.get('failed_data_backup') is not None:
                screenshot_val = extracted['failed_data_backup']
                reported_val = reported.get('failed_data_backup')
                
                if reported_val is not None and screenshot_val != reported_val:
                    issues.append(ValidationIssue(
                        sheet=analysis.sheet_name,
                        image_name=analysis.image_name,
                        severity='critical',
                        message=f'Screenshot shows {screenshot_val} failed data backups but cell reports {reported_val}',
                        screenshot_value=screenshot_val,
                        reported_value=reported_val
                    ))
            
            # Check failed log backup
            if extracted.get('failed_log_backup') is not None:
                screenshot_val = extracted['failed_log_backup']
                reported_val = reported.get('failed_log_backup')
                
                if reported_val is not None and screenshot_val != reported_val:
                    issues.append(ValidationIssue(
                        sheet=analysis.sheet_name,
                        image_name=analysis.image_name,
                        severity='critical',
                        message=f'Screenshot shows {screenshot_val} failed log backups but cell reports {reported_val}',
                        screenshot_value=screenshot_val,
                        reported_value=reported_val
                    ))
            
            # Check failed jobs
            if extracted.get('failed_jobs') is not None:
                screenshot_val = extracted['failed_jobs']
                reported_val = reported.get('failed_jobs')
                
                if reported_val is not None and screenshot_val != reported_val:
                    issues.append(ValidationIssue(
                        sheet=analysis.sheet_name,
                        image_name=analysis.image_name,
                        severity='critical',
                        message=f'Screenshot shows {screenshot_val} failed jobs but cell reports {reported_val}',
                        screenshot_value=screenshot_val,
                        reported_value=reported_val
                    ))
            
            # Check for error indicators when report shows all OK
            # Only flag if ALL failure metrics (backup + jobs) are 0
            reported_failures = (
                reported.get('failed_data_backup', 0) + 
                reported.get('failed_log_backup', 0) + 
                reported.get('failed_jobs', 0)
            )
            if extracted.get('has_errors') and reported_failures == 0:
                issues.append(ValidationIssue(
                    sheet=analysis.sheet_name,
                    image_name=analysis.image_name,
                    severity='warning',
                    message=f'Screenshot shows error indicators but no failures reported: {extracted.get("error_indicators", [])}',
                    screenshot_value=extracted.get('error_indicators'),
                    reported_value=0
                ))
        
        self.validation_issues = issues
        
        self.validation_issues = issues
        return issues
    
    def run_validation(self) -> tuple[list[ScreenshotAnalysis], list[ValidationIssue]]:
        """Run the complete validation workflow."""
        print("[SCREENSHOT] Extracting images from Excel...")
        images = self.extract_images_from_excel()
        
        if not images:
            print("[SCREENSHOT] No embedded images found")
            return [], []
        
        print(f"[SCREENSHOT] Found {len(images)} embedded images")
        
        if not self.use_azure:
            print("[SCREENSHOT] Azure OpenAI API not configured - skipping vision analysis")
            print("[SCREENSHOT] Set AZURE_OPENAI_API_KEY and ENDPOINT in .env")
            return [], []
        
        print("[SCREENSHOT] Analyzing images with Azure OpenAI (Structured)...")
        self.analyze_all_images()
        
        print("[SCREENSHOT] Validating against reported values...")
        self.validate_against_reports()
        
        return self.analyses, self.validation_issues


def main():
    """Test the screenshot validator."""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python screenshot_validator.py <excel_file>")
        sys.exit(1)
    
    validator = ScreenshotValidator(sys.argv[1])
    analyses, issues = validator.run_validation()
    
    print(f"\n[RESULTS] Analyzed {len(analyses)} screenshots")
    print(f"[RESULTS] Found {len(issues)} validation issues")
    
    for issue in issues:
        print(f"  [{issue.severity.upper()}] {issue.sheet}: {issue.message}")


if __name__ == "__main__":
    main()
