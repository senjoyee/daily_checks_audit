#!/usr/bin/env python3
"""
SAP Daily Monitoring Checks Audit Script

Validates daily monitoring reports to ensure:
- All checks were performed correctly
- Negative findings have proper justifications
- Numeric values are within acceptable thresholds

Supports customer-specific threshold configurations via JSON config files.
"""

import sys
import re
import json
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def detect_customer(filename: str) -> str | None:
    """Detect customer name from filename prefix."""
    filename_upper = Path(filename).stem.upper()
    
    # Direct prefix matches
    for prefix in ['TBS', 'BSW', 'COREX', 'SONOCO']:
        if filename_upper.startswith(prefix):
            return prefix
    
    # EVIOSYS files use SONOCO config (company renamed)
    if 'EVIOSYS' in filename_upper:
        return 'SONOCO'
    
    return None


def load_customer_config(customer: str, script_dir: Path) -> dict | None:
    """Load customer-specific config from configs directory."""
    config_path = script_dir.parent / 'configs' / f'{customer}_config.json'
    if config_path.exists():
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return None


def get_threshold(config: dict | None, metric: str, default: int) -> int:
    """Get threshold from config or return default."""
    if config and 'thresholds' in config:
        if metric in config['thresholds']:
            return config['thresholds'][metric].get('max', default)
    return default


def get_warning_threshold(config: dict | None, metric: str, default: int) -> int:
    """Get warning threshold from config or return default."""
    if config and 'thresholds' in config:
        if metric in config['thresholds']:
            return config['thresholds'][metric].get('warning', default)
    return default


# ============================================================================
# VALIDATION RULES CONFIGURATION
# ============================================================================
VALIDATION_RULES = {
    'response_time_smlg': {
        'threshold': 1000,
        'operator': '<',
        'severity': 'warning',
        'message': 'Response time exceeds 1000ms threshold'
    },
    'response_time_st03n': {
        'threshold': 1000,
        'operator': '<',
        'severity': 'warning',
        'message': 'Dialog/RFC response time exceeds 1000ms'
    },
    'dump_count_today': {
        'threshold': 50,
        'operator': '<',
        'severity': 'warning',
        'message': 'High number of ABAP dumps today'
    },
    'dump_count_yesterday': {
        'threshold': 100,
        'operator': '<',
        'severity': 'warning',
        'message': 'High number of ABAP dumps yesterday'
    },
    'failed_updates': {
        'threshold': 0,
        'operator': '==',
        'severity': 'critical',
        'message': 'Failed updates detected'
    },
    'trfc_errors': {
        'threshold': 0,
        'operator': '==',
        'severity': 'critical',
        'message': 'tRFC errors (CPICERR/SYSFAIL) detected'
    },
    'failed_jobs': {
        'threshold': 10,
        'operator': '<',
        'severity': 'warning',
        'message': 'High number of failed jobs'
    }
}

# Check patterns to identify row types
CHECK_PATTERNS = {
    'sm51': r'SM51|application server.*running',
    'sm50': r'SM50|SM66|work process',
    'smlg': r'SMLG|response time',
    'sm21': r'SM21|system log',
    'sm37': r'SM37|cancelled.*job|failed.*job',
    'sm12': r'SM12|old lock',
    'st22': r'ST22|abap.*dump',
    'dbacockpit': r'DBACOCKPIT|database.*performance',
    'sm13': r'SM13|update.*monitoring|failed update',
    'st02': r'ST02|buffer',
    'st03n': r'ST03N|workload.*monitoring',
    'spad': r'SPAD|spool',
    'sm58': r'SM58|trfc',
    'sost': r'SOST|failed.*email',
    'bop_cmc': r'CMC|server.*status',
    'nwa': r'NWA|system overview'
}


class AuditIssue:
    """Represents a single audit finding."""
    
    def __init__(self, sheet: str, row: int, check_type: str, 
                 severity: str, message: str, context: str = ""):
        self.sheet = sheet
        self.row = row
        self.check_type = check_type
        self.severity = severity  # 'critical', 'warning', 'info'
        self.message = message
        self.context = context
    
    def __repr__(self):
        return f"[{self.severity.upper()}] {self.sheet} Row {self.row}: {self.message}"


class DailyChecksAuditor:
    """Audits SAP daily monitoring Excel reports."""
    
    def __init__(self, excel_path: str, config: dict | None = None, customer: str | None = None):
        self.excel_path = Path(excel_path)
        self.workbook = None
        self.issues: list[AuditIssue] = []
        self.metadata: dict[str, dict] = {}
        self.config = config
        self.customer = customer
        self.screenshot_stats = {'analyzed': 0, 'issues': 0}
        
    def load_workbook(self) -> bool:
        """Load the Excel workbook."""
        try:
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            return True
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return False
    
    def extract_metadata(self, sheet) -> dict:
        """Extract header metadata from a sheet."""
        metadata = {
            'system_name': None,
            'date': None,
            'time': None,
            'performed_by': None
        }
        
        for row_idx in range(1, 6):
            row = [cell.value for cell in sheet[row_idx]]
            if row:
                label = str(row[0]).strip().lower() if row[0] else ""
                value = row[1] if len(row) > 1 else None
                
                if 'system name' in label:
                    metadata['system_name'] = value
                elif 'date' in label:
                    metadata['date'] = value
                elif 'time' in label:
                    metadata['time'] = value
                elif 'performed by' in label:
                    metadata['performed_by'] = value
        
        return metadata
    
    def identify_check_type(self, row_data: list) -> str | None:
        """Identify the type of check based on row content."""
        row_text = ' '.join(str(cell) for cell in row_data if cell).lower()
        
        for check_type, pattern in CHECK_PATTERNS.items():
            if re.search(pattern, row_text, re.IGNORECASE):
                return check_type
        return None
    
    def get_cell_value(self, row: list, col_idx: int) -> Any:
        """Safely get a cell value from a row."""
        if col_idx < len(row):
            return row[col_idx]
        return None
    
    def is_negative_response(self, row: list) -> bool:
        """Check if the row contains a negative (N) response."""
        # Check columns D, E for 'N' response (0-indexed: 3, 4)
        for col_idx in [3, 4]:
            val = self.get_cell_value(row, col_idx)
            if val and str(val).strip().upper() == 'N':
                return True
        return False
    
    def has_justification(self, row: list) -> tuple[bool, str]:
        """Check if a negative response has a justification in Status column (G)."""
        status = self.get_cell_value(row, 6)  # Column G (0-indexed: 6)
        
        if status:
            status_str = str(status).strip()
            # Filter out empty or whitespace-only content
            if status_str and status_str not in ['\xa0', ' ']:
                return True, status_str
        return False, ""
    
    def extract_numeric_value(self, row: list) -> tuple[float | None, int]:
        """Extract numeric value and its column index from a row."""
        # Check columns D, E, F for numeric values
        for col_idx in [3, 4, 5]:
            val = self.get_cell_value(row, col_idx)
            if val is not None:
                # Handle string numbers with formatting
                if isinstance(val, str):
                    # Remove thousand separators and handle decimal formats
                    cleaned = val.replace(',', '.').replace(' ', '')
                    try:
                        return float(cleaned), col_idx
                    except ValueError:
                        continue
                elif isinstance(val, (int, float)):
                    return float(val), col_idx
        return None, -1
    
    def audit_sheet(self, sheet_name: str) -> list[AuditIssue]:
        """Audit a single sheet and return issues found."""
        sheet = self.workbook[sheet_name]
        issues = []
        current_check_type = None
        
        self.metadata[sheet_name] = self.extract_metadata(sheet)
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=6, values_only=True), 6):
            row_list = list(row)
            
            # Skip empty rows
            if not any(cell for cell in row_list):
                continue
            
            # Identify check type from header rows
            check_type = self.identify_check_type(row_list)
            if check_type:
                current_check_type = check_type
            
            # Check for negative responses without justification
            if self.is_negative_response(row_list):
                has_just, justification = self.has_justification(row_list)
                
                if not has_just:
                    issues.append(AuditIssue(
                        sheet=sheet_name,
                        row=row_idx,
                        check_type=current_check_type or 'unknown',
                        severity='critical',
                        message='Negative (N) response without justification',
                        context=str(row_list[0:4]) if row_list else ""
                    ))
                else:
                    # Check if justification is substantive (not just a code or short text)
                    if len(justification) < 10:
                        issues.append(AuditIssue(
                            sheet=sheet_name,
                            row=row_idx,
                            check_type=current_check_type or 'unknown',
                            severity='warning',
                            message=f'Negative response has brief justification: "{justification}"',
                            context=str(row_list[0:4]) if row_list else ""
                        ))
            
            # Check numeric thresholds based on check type
            row_text = ' '.join(str(cell) for cell in row_list if cell).lower()
            numeric_val, _ = self.extract_numeric_value(row_list)
            
            if numeric_val is not None:
                # Response time checks (SMLG)
                resp_threshold = get_threshold(self.config, 'response_time_smlg', 1000)
                if 'resp time' in row_text and numeric_val > resp_threshold:
                    issues.append(AuditIssue(
                        sheet=sheet_name,
                        row=row_idx,
                        check_type='smlg',
                        severity='warning',
                        message=f'Response time {numeric_val}ms exceeds {resp_threshold}ms threshold',
                        context=str(row_list[1:4]) if row_list else ""
                    ))
                
                # Dump count checks
                dumps_today_threshold = get_threshold(self.config, 'dumps_today', 50)
                if 'dump' in row_text and 'today' in row_text and numeric_val > dumps_today_threshold:
                    issues.append(AuditIssue(
                        sheet=sheet_name,
                        row=row_idx,
                        check_type='st22',
                        severity='warning',
                        message=f'High dump count today: {int(numeric_val)} (threshold: {dumps_today_threshold})',
                        context=""
                    ))
                
                dumps_yesterday_threshold = get_threshold(self.config, 'dumps_yesterday', 100)
                if 'dump' in row_text and 'yesterday' in row_text and numeric_val > dumps_yesterday_threshold:
                    issues.append(AuditIssue(
                        sheet=sheet_name,
                        row=row_idx,
                        check_type='st22',
                        severity='warning',
                        message=f'High dump count yesterday: {int(numeric_val)} (threshold: {dumps_yesterday_threshold})',
                        context=""
                    ))
                
                # Failed updates
                if 'failed update' in row_text and numeric_val > 0:
                    issues.append(AuditIssue(
                        sheet=sheet_name,
                        row=row_idx,
                        check_type='sm13',
                        severity='critical',
                        message=f'Failed updates detected: {int(numeric_val)}',
                        context=""
                    ))
                
                # tRFC errors
                if ('cpicerr' in row_text or 'sysfail' in row_text) and numeric_val > 0:
                    issues.append(AuditIssue(
                        sheet=sheet_name,
                        row=row_idx,
                        check_type='sm58',
                        severity='critical',
                        message=f'tRFC errors detected: {int(numeric_val)}',
                        context=""
                    ))
                
                # Old locks
                if 'old lock' in row_text or 'number of old locks' in row_text:
                    if numeric_val > 0:
                        # Check if there's an explanation
                        has_just, _ = self.has_justification(row_list)
                        if not has_just:
                            issues.append(AuditIssue(
                                sheet=sheet_name,
                                row=row_idx,
                                check_type='sm12',
                                severity='warning',
                                message=f'Old locks present ({int(numeric_val)}) without explanation',
                                context=""
                            ))
        
        return issues
    
    def audit_all_sheets(self) -> None:
        """Audit all sheets in the workbook."""
        for sheet_name in self.workbook.sheetnames:
            sheet_issues = self.audit_sheet(sheet_name)
            self.issues.extend(sheet_issues)
    
    def generate_report(self) -> str:
        """Generate a markdown audit report."""
        report = []
        
        # Header
        report.append(f"# Audit Report - {self.excel_path.name}\n")
        report.append(f"**Generated**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        # Executive Summary
        critical_count = sum(1 for i in self.issues if i.severity == 'critical')
        warning_count = sum(1 for i in self.issues if i.severity == 'warning')
        
        report.append("## Executive Summary\n")
        report.append(f"- **Systems Checked**: {len(self.workbook.sheetnames)}")
        report.append(f"- **Total Issues**: {len(self.issues)}")
        report.append(f"- **Critical**: {critical_count} | **Warnings**: {warning_count}")
        if self.screenshot_stats['analyzed'] > 0:
            report.append(f"- **Screenshots Analyzed**: {self.screenshot_stats['analyzed']} (Issues: {self.screenshot_stats['issues']})")
        report.append("")
        
        if len(self.issues) == 0:
            report.append("> **All checks passed validation!**\n")
        
        # Metadata summary
        report.append("## Check Metadata\n")
        report.append("| System | Date | Time | Performed By |")
        report.append("|--------|------|------|--------------|")
        for sheet_name, meta in self.metadata.items():
            date_str = str(meta['date'])[:10] if meta['date'] else 'N/A'
            time_str = str(meta['time']) if meta['time'] else 'N/A'
            report.append(f"| {sheet_name} | {date_str} | {time_str} | {meta['performed_by'] or 'N/A'} |")
        report.append("")
        
        # Per-system breakdown
        report.append("## Per-System Findings\n")
        
        for sheet_name in self.workbook.sheetnames:
            sheet_issues = [i for i in self.issues if i.sheet == sheet_name]
            
            if sheet_issues:
                critical = [i for i in sheet_issues if i.severity == 'critical']
                warnings = [i for i in sheet_issues if i.severity == 'warning']
                
                status = "[CRITICAL]" if critical else "[WARNING]"
                report.append(f"### {status} {sheet_name}\n")
                report.append(f"**Issues Found**: {len(sheet_issues)} ({len(critical)} critical, {len(warnings)} warnings)\n")
                
                if critical:
                    report.append("#### Critical Issues\n")
                    for issue in critical:
                        report.append(f"- **Row {issue.row}** [{issue.check_type}]: {issue.message}")
                        if issue.context:
                            report.append(f"  - Context: `{issue.context[:80]}...`" if len(issue.context) > 80 else f"  - Context: `{issue.context}`")
                    report.append("")
                
                if warnings:
                    report.append("#### Warnings\n")
                    for issue in warnings:
                        report.append(f"- **Row {issue.row}** [{issue.check_type}]: {issue.message}")
                    report.append("")
            else:
                report.append(f"### [OK] {sheet_name}\n")
                report.append("All checks passed validation.\n")
        
        # Recommendations
        if self.issues:
            report.append("## Recommendations\n")
            
            if critical_count > 0:
                report.append("### Immediate Actions Required\n")
                report.append("1. Review all **critical** issues - these require immediate attention")
                report.append("2. Ensure negative responses have proper justifications with ticket numbers if applicable")
                report.append("3. Follow up with the team member who performed the checks\n")
            
            if warning_count > 0:
                report.append("### Follow-up Items\n")
                report.append("1. Review warning items for potential issues")
                report.append("2. Consider adjusting thresholds if warnings are expected behavior")
                report.append("3. Document any recurring patterns for process improvement\n")
        
        return '\n'.join(report)
    
    def save_report(self, output_path: str = None) -> str:
        """Save the audit report to a file."""
        if output_path is None:
            date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = self.excel_path.parent / f"audit_report_{date_str}.md"
        
        report_content = self.generate_report()
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(report_content)
        
        return str(output_path)


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("Usage: python audit_checks.py <excel_file_path>")
        print("Example: python audit_checks.py TBS_DAILY_MONITORING_20_JAN_2026.xlsx")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    
    if not Path(excel_path).exists():
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)
    
    print(f"[AUDIT] Auditing: {excel_path}")
    
    # Detect customer and load config
    script_dir = Path(__file__).parent
    customer = detect_customer(excel_path)
    config = None
    
    if customer:
        config = load_customer_config(customer, script_dir)
        if config:
            print(f"[CONFIG] Loaded {customer} customer configuration")
        else:
            print(f"[CONFIG] No config found for {customer}, using defaults")
    else:
        print("[CONFIG] Unknown customer format, using default thresholds")
    
    auditor = DailyChecksAuditor(excel_path, config=config, customer=customer)
    
    if not auditor.load_workbook():
        sys.exit(1)
    
    print(f"[INFO] Found {len(auditor.workbook.sheetnames)} system sheets")
    
    auditor.audit_all_sheets()
    
    # Screenshot validation (if images present)
    screenshot_issues = []
    try:
        from screenshot_validator import ScreenshotValidator
        validator = ScreenshotValidator(excel_path)
        analyses, screenshot_issues = validator.run_validation()
        
        # Add screenshot issues to the main issues list
        for issue in screenshot_issues:
            auditor.issues.append(AuditIssue(
                sheet=issue.sheet,
                row=0,  # Screenshots don't have row numbers
                check_type='screenshot_validation',
                severity=issue.severity,
                message=f"[SCREENSHOT] {issue.message}",
                context=f"Image: {issue.image_name}"
            ))
            
        auditor.screenshot_stats['analyzed'] = len(analyses)
        auditor.screenshot_stats['issues'] = len(screenshot_issues)
    except ImportError as e:
        print(f"[SCREENSHOT] Warning: Screenshot validator module could not be imported: {e}")
        # Debugging path issues
        # import sys
        # print(f"DEBUG: sys.path: {sys.path}")
    except Exception as e:
        print(f"[SCREENSHOT] Warning: Could not validate screenshots: {e}")
    
    # Print summary
    critical = sum(1 for i in auditor.issues if i.severity == 'critical')
    warnings = sum(1 for i in auditor.issues if i.severity == 'warning')
    screenshot_count = len(screenshot_issues)
    
    print(f"\n[RESULTS] Audit Results:")
    print(f"   [!] Critical: {critical}")
    print(f"   [?] Warnings: {warnings}")
    if auditor.screenshot_stats['analyzed'] > 0:
        print(f"   [IMG] Screenshots Analyzed: {auditor.screenshot_stats['analyzed']} (Issues: {screenshot_count})")
    print(f"   Total Issues: {len(auditor.issues)}")
    
    # Save report
    report_path = auditor.save_report()
    print(f"\n[OK] Report saved to: {report_path}")
    
    # Print report to console as well
    print("\n" + "="*60)
    print(auditor.generate_report())


if __name__ == "__main__":
    main()
